import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from io import BytesIO

st.title("Extract Key Ingredients 💋💄")

# Streamlit file uploader
uploaded_file = st.file_uploader("Upload a blank excel file with your ingredients in Column A", type=["xlsx"])

if uploaded_file:
    # Load workbook from uploaded file
    wb = load_workbook(uploaded_file)
    ws = wb.active

    # List of ingredients to search for
    ingredient_list = [
        "Algae", "Almond Oil", "Aloe Vera", "Alpha Arbutin", "Alpha Lipoic Acid", "Apricot Oil",
        "Argan Oil", "Azelaic Acid", "Baobab Oil & Powder", "Benzoyl Peroxide", "Biotin",
        "Buckthorn Oil", "Caffeine", "Cannabis Sativa Seed Oil", "Castor Oil", "Ceramides",
        "Centella Asiatica", "Chamomile", "Charcoal", "Citric Acid", "Clay", "Co-Enzyme Q10",
        "Cocoa Butter & Powder", "Coconut oil", "Collagen", "Copper", "DMAE", "Enzymes",
        "Eucalyptus Oil", "Ferulic Acid", "Ginseng", "Glycerin", "Glycolic Acid", "Grape Seed Oil",
        "Honey", "Hyaluronic Acid", "Jojoba Oil", "Lactic Acid", "Lanolin", "Lavender Oil",
        "Licorice Extract", "Mandelic Acid", "Marula Oil", "Micelles", "Milk Proteins", "Neroli Oil",
        "Niacinamide", "Omegas", "Panthenol", "Peptides", "Phytic Acid", "Purslane", "Resveratrol",
        "Retinol", "Rice", "Rose Oil & Rosewater", "Rosehip Oil & Rosehip", "Salicylic Acid",
        "Seaweed", "Shea Butter & Oil", "Silica", "SPF/Sunscreen Chemical", "SPF/Sunscreen Physical",
        "Squalane", "Sulphur", "Sunflower Seed Extract", "Tea", "Tea Tree Oil", "Vitamin C",
        "Vitamin E", "Witch Hazel"
    ]

    ingredient_list_lower = [ing.lower() for ing in ingredient_list]

    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def find_ingredients(cell_text):
        text = str(cell_text).lower()
        parts = re.split(r'[,\.;/|&•]+', text)
        parts = [p.strip() for p in parts if p.strip()]
        found = []
        for ing, ing_lower in zip(ingredient_list, ingredient_list_lower):
            if any(p == ing_lower for p in parts):
                found.append(ing)
                continue
            if re.search(r'\b' + re.escape(ing_lower) + r'\b', text):
                found.append(ing)
        return found

    # Process each cell in Column A
    for row in range(1, ws.max_row + 1):
        cell_value = ws[f"A{row}"].value
        if cell_value:
            found_ingredients = find_ingredients(cell_value)
            if found_ingredients:
                ws[f"B{row}"].value = "||".join(found_ingredients)
                ws[f"B{row}"].fill = highlight_fill

    # Save workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("Excel file processed successfully!")
    st.download_button(
        label="Download Processed Excel",
        data=output,
        file_name="processed_ingredients.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


